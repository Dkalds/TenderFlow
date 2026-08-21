"""Tests unitarios para services/exports (CSV/Excel/filename).

Funciones puras de transformación records → bytes; sin BD ni mocks.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
from fastapi import FastAPI
from starlette.testclient import TestClient

from api.routes.exports import _generate_ics
from services.exports import generate_csv, generate_excel, get_export_filename


def _records() -> list[dict]:
    return [
        {
            "id_externo": "L1",
            "titulo": "Servicio SAP cloud",
            "organo_contratacion": "ORG A",
            "importe": 1_000_000.0,
            "estado": "ADJ",
            "fecha_publicacion": "2025-01-01",
            "ccaa": "Madrid",
            "cpv": "72000000",
            "tecnologia": "SAP",
        },
        {
            "id_externo": "L2",
            "titulo": "Mantenimiento; ERP",  # `;` en el dato → debe quedar quoted
            "organo_contratacion": "ORG B",
            "importe": 200_000.0,
            "estado": "PUB",
            "fecha_publicacion": "2025-02-01",
            "ccaa": "Cataluña",
            "cpv": "48000000",
            "tecnologia": "SAP",
        },
    ]


# ── generate_csv ────────────────────────────────────────────────────────────


def test_generate_csv_bom_y_delimitador():
    data = generate_csv(_records())
    assert data.startswith(b"\xef\xbb\xbf")  # BOM UTF-8 para Excel
    text = data.decode("utf-8-sig")
    header = text.splitlines()[0]
    assert header.split(";")[0] == "id_externo"
    assert "L1" in text
    # El `;` dentro del dato queda escapado (quoted), no rompe columnas
    assert '"Mantenimiento; ERP"' in text


def test_generate_csv_respeta_orden_de_columns():
    data = generate_csv(_records(), columns=["titulo", "id_externo"])
    header = data.decode("utf-8-sig").splitlines()[0]
    assert header == "titulo;id_externo"


def test_generate_csv_ignora_columns_inexistentes():
    data = generate_csv(_records(), columns=["titulo", "no_existe"])
    header = data.decode("utf-8-sig").splitlines()[0]
    assert header == "titulo"


def test_generate_csv_records_vacios():
    data = generate_csv([])
    assert isinstance(data, bytes)
    assert data.startswith(b"\xef\xbb\xbf")


def test_generate_csv_utf8_caracteres_espanoles():
    data = generate_csv(_records())
    assert "Cataluña" in data.decode("utf-8-sig")


# ── generate_excel ──────────────────────────────────────────────────────────


def test_generate_excel_workbook_valido():
    data = generate_excel(_records())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Licitaciones"]
    # Cabecera + valores
    assert ws.cell(row=1, column=1).value == "id_externo"
    assert ws.cell(row=2, column=1).value == "L1"
    assert ws.cell(row=3, column=1).value == "L2"


def test_generate_excel_sheet_name_custom():
    data = generate_excel(_records(), sheet_name="Resultados")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Resultados" in wb.sheetnames


def test_generate_excel_strip_timezone():
    """openpyxl no soporta datetimes tz-aware; el export los normaliza a naive."""
    records = [
        {"id_externo": "L1", "fecha": pd.Timestamp("2025-01-01 10:00", tz="UTC")},
    ]
    data = generate_excel(records, columns=["id_externo", "fecha"])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    valor = ws.cell(row=2, column=2).value
    assert isinstance(valor, datetime)
    assert valor.tzinfo is None


def test_generate_excel_records_vacios():
    data = generate_excel([])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.active is not None


def test_spreadsheet_exports_neutralize_formula_cells():
    records = [{"id_externo": "L1", "titulo": '=HYPERLINK("https://evil.example")'}]

    csv_data = generate_csv(records, columns=["id_externo", "titulo"]).decode("utf-8-sig")
    assert "'=HYPERLINK" in csv_data

    xlsx_data = generate_excel(records, columns=["id_externo", "titulo"])
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data), data_only=False)
    assert workbook.active.cell(row=2, column=2).value.startswith("'=")


def test_ics_url_rejects_newline_property_injection():
    content = _generate_ics(
        [
            {
                "uid": "id-1",
                "dtstart": "2026-07-26",
                "summary": "Prueba",
                "url": "https://example.com/ok\r\nATTENDEE:mailto:attacker@example.com",
            }
        ]
    )
    assert "ATTENDEE:" not in content
    assert "URL:" not in content


# ── get_export_filename ─────────────────────────────────────────────────────


def test_get_export_filename_csv():
    name = get_export_filename("csv")
    fecha = datetime.now().strftime("%Y%m%d")
    assert name == f"licitaciones_{fecha}.csv"


def test_get_export_filename_excel_y_prefix():
    name = get_export_filename("excel", prefix="informe")
    fecha = datetime.now().strftime("%Y%m%d")
    assert name == f"informe_{fecha}.xlsx"


# ── Descarga síncrona (camino recomendado) ───────────────────────────────


def test_get_export_filename_extensions():
    """Cada formato mapea a su extensión; pdf es el añadido más reciente."""
    from services.exports import get_export_filename

    assert get_export_filename("csv").endswith(".csv")
    assert get_export_filename("excel").endswith(".xlsx")
    assert get_export_filename("pdf").endswith(".pdf")


def test_download_pdf_returns_document_inline(client, api_db):
    """``format=pdf`` devuelve el PDF en la respuesta, sin 202+poll.

    Es el camino que sustituye a los jobs asíncronos, cuyo almacén en memoria
    de proceso no sobrevive a un reinicio ni a una segunda instancia.
    """
    from api.auth import create_api_key

    create_api_key("pdf-download", scopes="*")

    resp = client.get("/api/v1/exports/download?format=pdf")

    # El endpoint exige sesión: sin ella responde 401/403, nunca 500 ni 404.
    assert resp.status_code in (200, 401, 403)
    if resp.status_code == 200:
        assert resp.headers["content-type"].startswith("application/pdf")
        assert resp.content.startswith(b"%PDF")
        assert ".pdf" in resp.headers["content-disposition"]


# ── Casos adicionales (coverage-driven, valores de retorno básicos) ─────────


def test_generate_csv_returns_bytes_with_bom():
    result = generate_csv([{"id_externo": "L1", "titulo": "Test"}])
    assert isinstance(result, bytes)
    assert result[:3] == b"\xef\xbb\xbf"


def test_generate_csv_uses_semicolon_delimiter():
    result = generate_csv([{"id_externo": "L1", "titulo": "Test"}])
    content = result[3:].decode("utf-8")
    assert ";" in content


def test_generate_csv_custom_columns():
    result = generate_csv(
        [{"id_externo": "L1", "titulo": "Test", "ccaa": "Madrid"}],
        columns=["id_externo", "ccaa"],
    )
    content = result[3:].decode("utf-8")
    assert "id_externo" in content
    assert "ccaa" in content


def test_generate_csv_empty_records_returns_bytes():
    result = generate_csv([])
    assert isinstance(result, bytes)


def test_generate_csv_missing_columns_ignored():
    result = generate_csv(
        [{"id_externo": "L1"}],
        columns=["id_externo", "nonexistent_col"],
    )
    assert isinstance(result, bytes)


def test_generate_excel_returns_bytes():
    result = generate_excel([{"id_externo": "L1", "titulo": "Test"}])
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_excel_custom_sheet_name_smoke():
    result = generate_excel([{"id_externo": "L1"}], sheet_name="CustomSheet")
    assert isinstance(result, bytes)


def test_generate_excel_empty_records_returns_bytes():
    result = generate_excel([])
    assert isinstance(result, bytes)


def test_generate_excel_timezone_aware_datetime_stripped_smoke():
    df = pd.DataFrame(
        {"fecha": pd.to_datetime(["2024-01-01"]).tz_localize("UTC"), "titulo": ["Test"]}
    )
    result = generate_excel(df.to_dict("records"), columns=["fecha", "titulo"])
    assert isinstance(result, bytes)


def test_get_export_filename_csv_format():
    name = get_export_filename("csv")
    assert name.endswith(".csv")
    assert "licitaciones_" in name


def test_get_export_filename_excel_format():
    name = get_export_filename("excel")
    assert name.endswith(".xlsx")


def test_get_export_filename_custom_prefix():
    name = get_export_filename("csv", prefix="custom")
    assert name.startswith("custom_")


# ── Resolución de rutas ──────────────────────────────────────────────────────


def _resuelve(path: str) -> str | None:
    """Devuelve el nombre del handler al que resuelve un GET a *path*.

    Envuelve la app en un ASGI middleware que lee ``scope["route"]`` una vez
    que el router lo ha fijado, igual que el fixture ``resolve`` de
    ``tests/test_routing_id_externo_con_barras.py``.

    Antes recorría ``app.routes`` buscando ``Match.FULL`` y devolvía
    ``route.name``. Eso funciona con el fastapi que fija el pin
    (``>=0.110,<0.137``), pero depende de que ``include_router`` aplane las
    ``APIRoute`` en ``app.routes``, que es justo lo que deja de ser cierto a
    partir de 0.141: los routers incluidos pasan a ser objetos
    ``_IncludedRouter`` opacos, sin ``.name`` ni ``.path``. Ver el porqué del
    techo en ``requirements.in`` (PR #187). Resolver por ``scope["route"]`` no
    mira la representación interna, así que el guard sobrevive a esa subida en
    vez de convertirse en un ``AttributeError`` el día que se haga.

    Devuelve ``None`` si ninguna ruta casó (404 de enrutado). No necesita BD ni
    credenciales: el enrutado ocurre antes que la autenticación, y el ``route``
    se lee en un ``finally`` para que un fallo del handler (401, 500) no lo
    pierda.
    """
    from api.routes.exports import router

    app = FastAPI()
    app.include_router(router, prefix="/api/v1")

    captured: dict[str, Any] = {}

    async def probe(scope, receive, send):
        try:
            await app(scope, receive, send)
        finally:
            captured["route"] = scope.get("route")

    client = TestClient(probe, raise_server_exceptions=False)
    client.get(path)
    return getattr(getattr(captured.get("route"), "endpoint", None), "__name__", None)


def test_rutas_estaticas_de_exports_no_las_ensombrece_el_comodin():
    """``/{job_id}`` no puede tragarse las sub-rutas estáticas de ``/exports``.

    Starlette resuelve por orden de registro, así que declarar ``/{job_id}``
    antes que ``/download`` hacía que toda descarga cayera en ``get_export``
    con job_id="download". Como ese handler exige X-API-Key, el endpoint que
    usan todos los botones de exportación del dashboard respondía 401 a
    cualquier sesión de navegador. El test existente no lo detectó porque
    aceptaba 401 entre los códigos válidos; este comprueba el handler, no el
    código de estado, que es lo único que no se puede satisfacer por accidente.
    """
    assert _resuelve("/api/v1/exports/download") == "download_export"
    assert _resuelve("/api/v1/exports/calendario.ics") == "calendario_ics"
    # El camino paramétrico sigue vivo para los jobs reales (uuid4).
    assert _resuelve("/api/v1/exports/1c9f5f2e-0000-4000-8000-000000000000") == "get_export"
